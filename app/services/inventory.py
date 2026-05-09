import io
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from app.services.auth import api_post, api_get, TENANT_URL, FACILITY_CODE


FACILITY_MAP = {
    "Emiza_B2C_BLR":    "Bangalore",
    "Emiza_B2C_GGN":    "Gurugram",
    "Emiza_B2C_Mumbai": "Mumbai",
    "Emiza_B2C_WB":     "West Bengal",
}

# Unicommerce requires updatedSinceInMinutes or itemTypeSKUs — never accepts an empty body.
# Hard limit confirmed from API error: "You can query for only one day snapshots" → max 1440 minutes.
_FULL_SNAPSHOT_MINUTES = 1_440


def fetch_facilities() -> list[str]:
    return list(FACILITY_MAP.keys())


def fetch_inventory(
    updated_since_minutes: Optional[int] = None,
) -> list[dict]:
    url            = f"{TENANT_URL}/services/rest/v1/inventory/inventorySnapshot/get"
    facility_codes = fetch_facilities()

    # Always use the caller-supplied window; fall back to full-snapshot sentinel.
    minutes = updated_since_minutes if updated_since_minutes is not None else _FULL_SNAPSHOT_MINUTES
    payload = {"updatedSinceInMinutes": minutes}

    all_records = []
    seen_keys   = set()

    for code in facility_codes:
        try:
            data = api_post(url, payload, facility=code)

            if data.get("successful"):
                records = (
                    data.get("inventorySnapshots")
                    or data.get("inventorySnapShotList")
                    or []
                )
                for r in records:
                    key = f"{r.get('itemTypeSKU')}_{code}"
                    if key not in seen_keys:
                        seen_keys.add(key)
                        r["facilityCode"]     = code
                        r["facilityLocation"] = FACILITY_MAP.get(code, code)
                        all_records.append(r)
                print(f"  ✔ Facility {code} ({FACILITY_MAP.get(code)}): {len(records)} SKUs")
            else:
                print(f"  ⚠ Facility {code}: {data.get('message')} | errors: {data.get('errors')}")

        except Exception as e:
            print(f"  ⚠ Facility {code} failed: {e}")
            continue

    print(f"Total records across all facilities: {len(all_records)}")
    return all_records


def build_inventory_excel(records: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory Snapshot"

    header_font  = Font(bold=True, color="FFFFFF", name="Arial")
    header_fill  = PatternFill("solid", start_color="1F4E79")
    center_align = Alignment(horizontal="center", vertical="center")

    headers = [
        "SKU Code",
        "Facility Code",
        "Location",
        "Sellable Qty",
        "Open Sale",
        "Open Purchase",
        "Putaway Pending",
        "Blocked Qty",
        "Pending Stock Transfer",
        "Vendor Inventory",
        "Virtual Inventory",
        "Pending Assessment",
        "Bad Inventory",
        "Inventory Not Synced",
        "Batch Recall Qty",
    ]

    for col, h in enumerate(headers, start=1):
        cell           = ws.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
    ws.row_dimensions[1].height = 22

    for row_idx, item in enumerate(records, start=2):
        ws.cell(row=row_idx, column=1,  value=item.get("itemTypeSKU"))
        ws.cell(row=row_idx, column=2,  value=item.get("facilityCode"))
        ws.cell(row=row_idx, column=3,  value=item.get("facilityLocation"))
        ws.cell(row=row_idx, column=4,  value=item.get("inventory", 0))
        ws.cell(row=row_idx, column=5,  value=item.get("openSale", 0))
        ws.cell(row=row_idx, column=6,  value=item.get("openPurchase", 0))
        ws.cell(row=row_idx, column=7,  value=item.get("putawayPending", 0))
        ws.cell(row=row_idx, column=8,  value=item.get("inventoryBlocked", 0))
        ws.cell(row=row_idx, column=9,  value=item.get("pendingStockTransfer", 0))
        ws.cell(row=row_idx, column=10, value=item.get("vendorInventory", 0))
        ws.cell(row=row_idx, column=11, value=item.get("virtualInventory", 0))
        ws.cell(row=row_idx, column=12, value=item.get("pendingInventoryAssessment", 0))
        ws.cell(row=row_idx, column=13, value=item.get("badInventory", 0))
        ws.cell(row=row_idx, column=14, value=item.get("inventoryNotSynced", 0))
        ws.cell(row=row_idx, column=15, value=item.get("batchRecallQuantity", 0))

        fill_color = "EBF3FB" if row_idx % 2 == 0 else "FFFFFF"
        row_fill   = PatternFill("solid", start_color=fill_color)
        for col in range(1, 16):
            ws.cell(row=row_idx, column=col).fill = row_fill

    col_widths = [16, 18, 14, 12, 10, 14, 14, 12, 22, 16, 16, 18, 14, 18, 16]
    for col, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    last = len(records) + 2
    total_fill = PatternFill("solid", start_color="D6E4F0")
    total_font = Font(bold=True, name="Arial")
    ws.cell(row=last, column=1, value="TOTAL").font = total_font

    numeric_cols = {
        4: "D", 5: "E", 6: "F", 7: "G", 8: "H",
        9: "I", 10: "J", 11: "K", 12: "L", 13: "M",
        14: "N", 15: "O"
    }
    for col_num, col_letter in numeric_cols.items():
        ws.cell(row=last, column=col_num, value=f"=SUM({col_letter}2:{col_letter}{last-1})")
        ws.cell(row=last, column=col_num).font = total_font

    for col in range(1, 16):
        ws.cell(row=last, column=col).fill = total_fill

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()