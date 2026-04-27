import io
from datetime import datetime
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from app.services.auth import api_post, TENANT_URL


# ── EXCEL PARSING ─────────────────────────────────────────────────────────────

def parse_po_excel(file_bytes: bytes) -> tuple[list[dict], list[str]]:
    """
    Expects two sheets:
      'PO Header' — one row per PO
      'PO Items'  — one row per line item, linked by purchaseOrderCode
    Returns (po_list, skipped_codes)
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    header_ws   = wb["PO Header"]
    header_cols = [cell.value for cell in header_ws[1]]

    po_map: dict[str, dict] = {}
    for row in header_ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        po      = dict(zip(header_cols, row))
        po_code = str(po["purchaseOrderCode"]).strip()
        po_map[po_code] = {
            "purchaseOrderCode":             po_code,
            "type":                          po.get("type", "Manual"),
            "vendorCode":                    str(po.get("vendorCode", "")).strip(),
            "vendorAgreementName":           str(po.get("vendorAgreementName", "")).strip() or None,
            "currencyCode":                  str(po.get("currencyCode", "INR")).strip(),
            "deliveryDate":                  _fmt_date(po.get("deliveryDate")),
            "expiryDate":                    _fmt_date(po.get("expiryDate")),
            "logisticCharges":               float(po.get("logisticCharges") or 0),
            "logisticChargesDivisionMethod": str(po.get("logisticChargesDivisionMethod", "")).strip() or None,
            "purchaseOrderItems":            [],
        }

    items_ws   = wb["PO Items"]
    items_cols = [cell.value for cell in items_ws[1]]

    skipped: list[str] = []
    for row in items_ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        item    = dict(zip(items_cols, row))
        po_code = str(item["purchaseOrderCode"]).strip()
        if po_code not in po_map:
            skipped.append(po_code)
            continue
        po_map[po_code]["purchaseOrderItems"].append({
            "itemSKU":            str(item.get("itemSKU", "")).strip(),
            "quantity":           int(item.get("quantity") or 0),
            "unitPrice":          float(item.get("unitPrice") or 0),
            "maxRetailPrice":     float(item.get("maxRetailPrice") or 0),
            "discount":           float(item.get("discount") or 0),
            "discountPercentage": float(item.get("discountPercentage") or 0),
            "taxTypeCode":        str(item.get("taxTypeCode", "")).strip() or None,
        })

    return list(po_map.values()), skipped


def _fmt_date(val) -> Optional[str]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%dT%H:%M:%S.000Z")
    return str(val).strip()


# ── API CALLS ─────────────────────────────────────────────────────────────────

def create_purchase_orders(po_list: list[dict], approve: bool = False) -> list[dict]:
    endpoint = "createAndApprove" if approve else "create"
    url      = f"{TENANT_URL}/services/rest/v1/purchase/purchaseOrder/{endpoint}"

    results = []
    for po in po_list:
        po_code = po["purchaseOrderCode"]
        if not po["purchaseOrderItems"]:
            results.append({"poCode": po_code, "status": "SKIPPED", "message": "No items found"})
            continue
        try:
            data = api_post(url, po)
            if data.get("successful"):
                results.append({"poCode": po_code, "status": "SUCCESS", "message": data.get("message", "")})
            else:
                errors = data.get("errors", [])
                msg    = errors[0].get("message") if errors else data.get("message", "Unknown error")
                results.append({"poCode": po_code, "status": "FAILED", "message": msg})
        except Exception as e:
            results.append({"poCode": po_code, "status": "ERROR", "message": str(e)})

    return results


# ── RESULTS EXCEL ─────────────────────────────────────────────────────────────

def build_results_excel(results: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Upload Results"

    header_font = Font(bold=True, color="FFFFFF", name="Arial")
    header_fill = PatternFill("solid", start_color="1F4E79")

    for col, h in enumerate(["PO Code", "Status", "Message"], start=1):
        cell           = ws.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")

    fill_map = {
        "SUCCESS": PatternFill("solid", start_color="C6EFCE"),
        "FAILED":  PatternFill("solid", start_color="FFC7CE"),
        "ERROR":   PatternFill("solid", start_color="FFC7CE"),
        "SKIPPED": PatternFill("solid", start_color="FFEB9C"),
    }

    for i, r in enumerate(results, start=2):
        ws.cell(row=i, column=1, value=r["poCode"])
        ws.cell(row=i, column=2, value=r["status"])
        ws.cell(row=i, column=3, value=r["message"])
        fill = fill_map.get(r["status"], fill_map["SKIPPED"])
        for col in range(1, 4):
            ws.cell(row=i, column=col).fill = fill

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 50

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── PO TEMPLATE ───────────────────────────────────────────────────────────────

def build_po_template() -> bytes:
    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color="FFFFFF", name="Arial")
    header_fill = PatternFill("solid", start_color="1F4E79")
    center      = Alignment(horizontal="center")

    ws1       = wb.active
    ws1.title = "PO Header"
    po_cols   = [
        "purchaseOrderCode", "type", "vendorCode", "vendorAgreementName",
        "currencyCode", "deliveryDate", "expiryDate",
        "logisticCharges", "logisticChargesDivisionMethod",
    ]
    for col, h in enumerate(po_cols, start=1):
        cell           = ws1.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        ws1.column_dimensions[cell.column_letter].width = 26
    ws1.append(["PO-2024-001", "Manual", "VENDOR001", "", "INR", "2024-12-31", "2024-12-31", 0, ""])

    ws2       = wb.create_sheet("PO Items")
    item_cols = [
        "purchaseOrderCode", "itemSKU", "quantity", "unitPrice",
        "maxRetailPrice", "discount", "discountPercentage", "taxTypeCode",
    ]
    for col, h in enumerate(item_cols, start=1):
        cell           = ws2.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        ws2.column_dimensions[cell.column_letter].width = 22
    ws2.append(["PO-2024-001", "SKU001", 10, 500.00, 999.00, 0, 0, "GST18"])
    ws2.append(["PO-2024-001", "SKU002",  5, 300.00, 599.00, 0, 0, "GST12"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
