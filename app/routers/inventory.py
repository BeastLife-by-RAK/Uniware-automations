#OLD CODE 
# import io
# from typing import Optional
# from fastapi import APIRouter, Query, HTTPException
# from fastapi.responses import StreamingResponse, JSONResponse

# from app.services.inventory import fetch_inventory, build_inventory_excel, fetch_facilities
# from app.services.sheets import push_inventory_to_sheets

# router = APIRouter(prefix="/inventory", tags=["Inventory"])


# @router.get("/facilities")
# def list_facilities():
#     """List all facility codes."""
#     try:
#         codes = fetch_facilities()
#         return JSONResponse(content={"count": len(codes), "facilities": codes})
#     except Exception as e:
#         raise HTTPException(status_code=502, detail=str(e))


# @router.get("/fetch")
# def get_inventory(
#     format: str = Query("json", description="Response format: json or excel"),
#     updated_since_minutes: Optional[int] = Query(
#         None,
#         description="Only SKUs updated in last N minutes. Omit for full snapshot."
#     ),
#     skus: Optional[str] = Query(
#         None,
#         description="Comma-separated SKU codes e.g. SKU001,SKU002"
#     ),
# ):
#     """
#     Fetch inventory snapshot from Unicommerce across all facilities.

#     - **format=json** → raw JSON (default)
#     - **format=excel** → downloadable .xlsx file
#     - **updated_since_minutes** → SKUs updated in last N minutes. Omit for full snapshot.
#     - **skus** → filter to specific SKUs
#     """
#     sku_list = [s.strip() for s in skus.split(",")] if skus else None

#     try:
#         records = fetch_inventory(
#             updated_since_minutes=updated_since_minutes,
#             sku_list=sku_list,
#         )
#     except Exception as e:
#         raise HTTPException(status_code=502, detail=str(e))

#     if format == "excel":
#         return StreamingResponse(
#             io.BytesIO(build_inventory_excel(records)),
#             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#             headers={"Content-Disposition": "attachment; filename=inventory_snapshot.xlsx"},
#         )

#     return JSONResponse(content={"count": len(records), "records": records})


# @router.post("/push-to-sheets")
# def push_to_sheets(
#     updated_since_minutes: Optional[int] = Query(
#         None,
#         description="Only SKUs updated in last N minutes. Omit for full snapshot."
#     ),
# ):
#     """
#     Fetch inventory from Unicommerce and push directly to Google Sheets.
#     Creates a new dated tab per facility e.g. GGN-2026-04-27.
#     Designed to be triggered by Cloud Scheduler.
#     """
#     try:
#         records = fetch_inventory(updated_since_minutes=updated_since_minutes)
#     except Exception as e:
#         raise HTTPException(status_code=502, detail=f"Unicommerce fetch failed: {e}")

#     if not records:
#         return JSONResponse(content={"message": "No records returned", "tabs_created": {}})

#     try:
#         summary = push_inventory_to_sheets(records)
#     except Exception as e:
#         raise HTTPException(status_code=502, detail=f"Google Sheets write failed: {e}")

#     return JSONResponse(content={
#         "message": "Successfully pushed to Google Sheets",
#         "tabs_created": summary,
#         "total_records": len(records),
#     })

# #additional function for testing 
# @router.get("/audit")
# def audit_missing_skus():
#     """
#     Compare catalog SKUs vs inventory snapshot SKUs.
#     Returns which SKUs are missing from the inventory response.
#     """
#     from app.services.auth import api_post, TENANT_URL

#     # Step 1 — get ALL SKUs from catalog
#     search_url = f"{TENANT_URL}/services/rest/v1/product/itemType/search"
#     catalog_skus = set()
#     start = 0

#     while True:
#         payload = {
#             "searchOptions": {
#                 "displayLength": 500,
#                 "displayStart": start,
#                 "getCount": True
#             }
#         }
#         data = api_post(search_url, payload)
#         elements = data.get("elements", [])
#         for item in elements:
#             if item.get("skuCode") and item.get("enabled", True):
#                 catalog_skus.add(item["skuCode"])
#         start += len(elements)
#         if not elements or start >= data.get("totalRecords", 0):
#             break

#     # Step 2 — get SKUs from inventory snapshot (current method)
#     inv_url = f"{TENANT_URL}/services/rest/v1/inventory/inventorySnapshot/get"
#     inventory_skus = set()

#     for code in fetch_facilities():
#         try:
#             data = api_post(inv_url, {"updatedSinceInMinutes": 1440}, facility=code)
#             for r in data.get("inventorySnapshots") or data.get("inventorySnapShotList") or []:
#                 inventory_skus.add(r.get("itemTypeSKU"))
#         except:
#             pass

#     # Step 3 — compare
#     missing = sorted(catalog_skus - inventory_skus)

#     return JSONResponse(content={
#         "catalog_total":    len(catalog_skus),
#         "inventory_total":  len(inventory_skus),
#         "missing_count":    len(missing),
#         "missing_skus":     missing,
#     })
#NEW CODE 
import io
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from app.services.auth import api_post, api_get, TENANT_URL, FACILITY_CODE


FACILITY_MAP = {
    "Emiza_B2C_BLR":    "Bangalore",
    "Emiza_B2C_GGN":    "Gurugram",
    "Emiza_B2C_Mumbai": "Mumbai",
    "Emiza_B2C_WB":     "West Bengal",
}

_SKU_PAGE_SIZE  = 500
_INV_CHUNK_SIZE = 500


def fetch_facilities() -> list[str]:
    return list(FACILITY_MAP.keys())


def fetch_all_sku_codes() -> list[str]:
    url   = f"{TENANT_URL}/services/rest/v1/product/itemType/search"
    start = 0
    skus  = []
    while True:
        payload = {
            "searchOptions": {
                "displayLength": _SKU_PAGE_SIZE,
                "displayStart":  start,
                "getCount":      True,
            }
        }
        data = api_post(url, payload)
        if not data.get("successful"):
            break
        elements = data.get("elements", [])
        for item in elements:
            sku = item.get("skuCode")
            if sku and item.get("enabled", True):
                skus.append(sku)
        start += len(elements)
        if not elements or start >= data.get("totalRecords", 0):
            break
    return skus


def fetch_inventory(
    updated_since_minutes: Optional[int] = None,
    sku_list: Optional[list[str]] = None,
) -> list[dict]:
    url            = f"{TENANT_URL}/services/rest/v1/inventory/inventorySnapshot/get"
    facility_codes = fetch_facilities()
    skus           = sku_list if sku_list else fetch_all_sku_codes()

    all_records = []
    seen_keys   = set()

    for code in facility_codes:
        try:
            for i in range(0, len(skus), _INV_CHUNK_SIZE):
                chunk = skus[i : i + _INV_CHUNK_SIZE]
                data  = api_post(url, {"itemTypeSKUs": chunk}, facility=code)
                if data.get("successful"):
                    records = (
                        data.get("inventorySnapshots")
                        or data.get("inventorySnapShotList")
                        or []
                    )
                    for r in records:
                        key = f"{r.get('itemTypeSKU')}_{code}"
                        if key not in seen_keys:
                            seen_keys.add(key)
                            r["facilityCode"]     = code
                            r["facilityLocation"] = FACILITY_MAP.get(code, code)
                            all_records.append(r)
                    print(f"  ✔ Facility {code} chunk {i // _INV_CHUNK_SIZE + 1}: {len(records)} SKUs")
                else:
                    print(f"  ⚠ Facility {code}: {data.get('message')} | errors: {data.get('errors')}")
        except Exception as e:
            print(f"  ⚠ Facility {code} failed: {e}")
            continue

    print(f"Total records across all facilities: {len(all_records)}")
    return all_records


def build_inventory_excel(records: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory Snapshot"

    header_font  = Font(bold=True, color="FFFFFF", name="Arial")
    header_fill  = PatternFill("solid", start_color="1F4E79")
    center_align = Alignment(horizontal="center", vertical="center")

    headers = [
        "SKU Code",
        "Facility Code",
        "Location",
        "Sellable Qty",
        "Open Sale",
        "Open Purchase",
        "Putaway Pending",
        "Blocked Qty",
        "Pending Stock Transfer",
        "Vendor Inventory",
        "Virtual Inventory",
        "Pending Assessment",
        "Bad Inventory",
        "Inventory Not Synced",
        "Batch Recall Qty",
    ]

    for col, h in enumerate(headers, start=1):
        cell           = ws.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
    ws.row_dimensions[1].height = 22

    for row_idx, item in enumerate(records, start=2):
        ws.cell(row=row_idx, column=1,  value=item.get("itemTypeSKU"))
        ws.cell(row=row_idx, column=2,  value=item.get("facilityCode"))
        ws.cell(row=row_idx, column=3,  value=item.get("facilityLocation"))
        ws.cell(row=row_idx, column=4,  value=item.get("inventory", 0))
        ws.cell(row=row_idx, column=5,  value=item.get("openSale", 0))
        ws.cell(row=row_idx, column=6,  value=item.get("openPurchase", 0))
        ws.cell(row=row_idx, column=7,  value=item.get("putawayPending", 0))
        ws.cell(row=row_idx, column=8,  value=item.get("inventoryBlocked", 0))
        ws.cell(row=row_idx, column=9,  value=item.get("pendingStockTransfer", 0))
        ws.cell(row=row_idx, column=10, value=item.get("vendorInventory", 0))
        ws.cell(row=row_idx, column=11, value=item.get("virtualInventory", 0))
        ws.cell(row=row_idx, column=12, value=item.get("pendingInventoryAssessment", 0))
        ws.cell(row=row_idx, column=13, value=item.get("badInventory", 0))
        ws.cell(row=row_idx, column=14, value=item.get("inventoryNotSynced", 0))
        ws.cell(row=row_idx, column=15, value=item.get("batchRecallQuantity", 0))

        fill_color = "EBF3FB" if row_idx % 2 == 0 else "FFFFFF"
        row_fill   = PatternFill("solid", start_color=fill_color)
        for col in range(1, 16):
            ws.cell(row=row_idx, column=col).fill = row_fill

    col_widths = [16, 18, 14, 12, 10, 14, 14, 12, 22, 16, 16, 18, 14, 18, 16]
    for col, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    last       = len(records) + 2
    total_fill = PatternFill("solid", start_color="D6E4F0")
    total_font = Font(bold=True, name="Arial")
    ws.cell(row=last, column=1, value="TOTAL").font = total_font

    numeric_cols = {
        4: "D", 5: "E", 6: "F", 7: "G", 8: "H",
        9: "I", 10: "J", 11: "K", 12: "L", 13: "M",
        14: "N", 15: "O",
    }
    for col_num, col_letter in numeric_cols.items():
        ws.cell(row=last, column=col_num, value=f"=SUM({col_letter}2:{col_letter}{last-1})")
        ws.cell(row=last, column=col_num).font = total_font

    for col in range(1, 16):
        ws.cell(row=last, column=col).fill = total_fill

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
